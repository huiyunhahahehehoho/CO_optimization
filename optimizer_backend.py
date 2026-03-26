from __future__ import annotations

import math
import shutil
import time
import warnings
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

# Optional solvers
try:
    import gurobipy as gp
    from gurobipy import GRB
    GUROBI_AVAILABLE = True
except Exception:
    gp = None
    GRB = None
    GUROBI_AVAILABLE = False

try:
    from ortools.constraint_solver import pywrapcp, routing_enums_pb2
    ORTOOLS_AVAILABLE = True
except Exception:
    pywrapcp = None
    routing_enums_pb2 = None
    ORTOOLS_AVAILABLE = False


FULL_ENUMERATION_LABEL = "Enumeration"

HIGH_COUNT_ALGORITHMS = [
    "Gurobi Exact",
    "OR-Tools TSP",
    "Ant Colony Optimization",
    "Christofides Algorithm",
    "2-opt / 3-opt",
    "Lin-Kernighan",
    "Simulated Annealing",
    "Tabu Search",
    "Cheapest Insertion",
    "TSP Heuristic",
]


TIME_LIMIT_SECONDS: Optional[int] = None


def normalize_gurobi_license_config(gurobi_license_config: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not gurobi_license_config:
        return None

    access_id = str(
        gurobi_license_config.get("WLSACCESSID")
        or gurobi_license_config.get("WLSAccessID")
        or ""
    ).strip()
    secret = str(
        gurobi_license_config.get("WLSSECRET")
        or gurobi_license_config.get("WLSSecret")
        or ""
    ).strip()
    license_id_raw = str(
        gurobi_license_config.get("LICENSEID")
        or gurobi_license_config.get("LicenseID")
        or ""
    ).strip()

    if not access_id and not secret and not license_id_raw:
        return None

    missing = []
    if not access_id:
        missing.append("WLSACCESSID")
    if not secret:
        missing.append("WLSSECRET")
    if not license_id_raw:
        missing.append("LICENSEID")
    if missing:
        raise ValueError("Please provide all Gurobi WLS fields: " + ", ".join(missing))

    license_id: Any = int(license_id_raw) if license_id_raw.isdigit() else license_id_raw
    return {
        "WLSAccessID": access_id,
        "WLSSecret": secret,
        "LicenseID": license_id,
    }


def create_gurobi_env(gurobi_license_config: Optional[Dict[str, Any]]):
    normalized = normalize_gurobi_license_config(gurobi_license_config)
    if normalized is None:
        return None

    params = {"OutputFlag": 0}
    params.update(normalized)
    return gp.Env(params=params)


def normalize_loop_type(value: Any) -> str:
    if value is None:
        return "open"
    text = str(value).strip().lower()
    if text in {"closed", "close", "closed loop", "closed-loop"}:
        return "closed"
    return "open"


def parse_time_limit_hours(value: Any) -> Optional[float]:
    if value in (None, "", "none", "None"):
        return None
    return float(value)


def refresh_time_limit_seconds(time_limit_hours: Any) -> None:
    global TIME_LIMIT_SECONDS
    parsed = parse_time_limit_hours(time_limit_hours)
    if parsed is None:
        TIME_LIMIT_SECONDS = None
    else:
        TIME_LIMIT_SECONDS = max(1, int(parsed * 3600))


def now() -> float:
    return time.time()


def elapsed(start_time: float) -> float:
    return time.time() - start_time


def within_limit(start_time: Optional[float]) -> bool:
    if start_time is None or TIME_LIMIT_SECONDS is None:
        return True
    return elapsed(start_time) < TIME_LIMIT_SECONDS


def remaining_time_seconds(start_time: Optional[float]) -> Optional[float]:
    if start_time is None or TIME_LIMIT_SECONDS is None:
        return None
    return max(0.0, TIME_LIMIT_SECONDS - elapsed(start_time))


def get_algorithm_options(problem_type: str, selected_count: int) -> List[str]:
    if 1 <= selected_count <= 15:
        return [FULL_ENUMERATION_LABEL]
    if selected_count > 15:
        return HIGH_COUNT_ALGORITHMS[:]
    return []


@lru_cache(maxsize=8)
def _load_co_matrix(matrix_path_str: str) -> pd.DataFrame:
    matrix_path = Path(matrix_path_str)
    if not matrix_path.exists():
        raise FileNotFoundError(f"CO Matrix file not found: {matrix_path}")
    co_df = pd.read_excel(matrix_path, sheet_name="CO Matrix", index_col=0)
    if co_df.empty:
        raise ValueError("The 'CO Matrix' sheet is empty.")
    return co_df


@lru_cache(maxsize=8)
def _load_product_maps(matrix_path_str: str) -> Tuple[List[str], Dict[str, Any], Dict[str, str]]:
    co_df = _load_co_matrix(matrix_path_str)
    actual_labels = list(co_df.index)
    display_products = [f"Product {i:03d}" for i in range(1, len(actual_labels) + 1)]
    display_to_actual = dict(zip(display_products, actual_labels))
    actual_to_display = {str(actual): display for display, actual in display_to_actual.items()}
    return display_products, display_to_actual, actual_to_display


def get_available_display_products(matrix_path: Path | str) -> List[str]:
    display_products, _, _ = _load_product_maps(str(Path(matrix_path).resolve()))
    return display_products


def resolve_display_products(selected_display_products: List[str], matrix_path: Path | str) -> List[Any]:
    _, display_to_actual, _ = _load_product_maps(str(Path(matrix_path).resolve()))
    missing = [name for name in selected_display_products if name not in display_to_actual]
    if missing:
        raise ValueError(f"These selected products are not available in the CO Matrix: {missing}")
    return [display_to_actual[name] for name in selected_display_products]


def get_cost_submatrix(selected_display_products: List[str], matrix_path: Path | str) -> Tuple[List[str], np.ndarray, List[Any]]:
    matrix_path = str(Path(matrix_path).resolve())
    co_df = _load_co_matrix(matrix_path)
    actual_labels = resolve_display_products(selected_display_products, matrix_path)
    sub_df = co_df.loc[actual_labels, actual_labels].copy()
    sub_df = sub_df.apply(pd.to_numeric, errors="coerce")
    if sub_df.isna().any().any():
        raise ValueError("Found non-numeric or blank values inside the selected CO Matrix area.")
    matrix = sub_df.to_numpy(dtype=float)
    return selected_display_products[:], matrix, actual_labels


def route_cost(route: List[int], cost_matrix: np.ndarray, closed_loop: bool) -> float:
    if route is None or len(route) == 0:
        return math.inf
    total = 0.0
    for i in range(len(route) - 1):
        total += float(cost_matrix[route[i], route[i + 1]])
    if closed_loop and len(route) > 1:
        total += float(cost_matrix[route[-1], route[0]])
    return float(total)


def route_to_display_sequence(route: Optional[List[int]], display_labels: List[str], closed_loop: bool = False) -> List[str]:
    if route is None:
        return []
    seq = [display_labels[i] for i in route]
    if closed_loop and seq:
        return seq + [seq[0]]
    return seq


def nearest_neighbor_route(cost_matrix: np.ndarray, start: int = 0) -> List[int]:
    n = len(cost_matrix)
    unvisited = set(range(n))
    route = [start]
    unvisited.remove(start)
    current = start
    while unvisited:
        nxt = min(unvisited, key=lambda j: cost_matrix[current, j])
        route.append(nxt)
        unvisited.remove(nxt)
        current = nxt
    return route


def two_opt(route: List[int], cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None) -> List[int]:
    best = route[:]
    best_cost = route_cost(best, cost_matrix, closed_loop)
    improved = True
    n = len(route)

    while improved:
        improved = False
        for i in range(1, n - 1):
            for j in range(i + 1, n):
                if t0 is not None and not within_limit(t0):
                    return best
                new_route = best[:]
                new_route[i:j] = reversed(new_route[i:j])
                new_cost = route_cost(new_route, cost_matrix, closed_loop)
                if new_cost + 1e-12 < best_cost:
                    best = new_route
                    best_cost = new_cost
                    improved = True
        route = best
    return best


def three_opt_simple(route: List[int], cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None) -> List[int]:
    best = route[:]
    best_cost = route_cost(best, cost_matrix, closed_loop)
    n = len(best)

    improved = True
    while improved:
        improved = False
        for i in range(1, n - 4):
            for j in range(i + 1, n - 2):
                for k in range(j + 1, n):
                    if t0 is not None and not within_limit(t0):
                        return best

                    a = best[:i]
                    b = best[i:j]
                    c = best[j:k]
                    d = best[k:]

                    candidates = [
                        a + b[::-1] + c + d,
                        a + b + c[::-1] + d,
                        a + c + b + d,
                        a + c[::-1] + b + d,
                        a + c + b[::-1] + d,
                        a + b[::-1] + c[::-1] + d,
                    ]

                    for cand in candidates:
                        cand_cost = route_cost(cand, cost_matrix, closed_loop)
                        if cand_cost + 1e-12 < best_cost:
                            best = cand
                            best_cost = cand_cost
                            improved = True
                            break
                    if improved:
                        break
                if improved:
                    break
            if improved:
                break
    return best


def symmetrize_matrix(cost_matrix: np.ndarray, mode: str = "avg") -> np.ndarray:
    a = np.array(cost_matrix, dtype=float)
    if mode == "min":
        return np.minimum(a, a.T)
    if mode == "max":
        return np.maximum(a, a.T)
    return 0.5 * (a + a.T)


# ------------------------------
# Exact methods
# ------------------------------
def solve_full_enumeration(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    n = len(cost_matrix)

    if n == 0:
        return [], 0.0, "OPTIMAL"
    if n == 1:
        return [0], 0.0, "OPTIMAL"

    best_route = None
    best_cost = math.inf
    checked = 0

    greedy_route, greedy_cost, _ = solve_greedy(cost_matrix, closed_loop, t0)
    if greedy_route is not None and greedy_cost is not None and np.isfinite(greedy_cost):
        best_route = greedy_route[:]
        best_cost = float(greedy_cost)

    def dfs(path, unvisited, current_cost):
        nonlocal best_route, best_cost, checked

        if t0 is not None and not within_limit(t0):
            return "TIME_LIMIT"

        if current_cost >= best_cost - 1e-12:
            return None

        if not unvisited:
            total_cost = current_cost
            if closed_loop and len(path) > 1:
                total_cost += float(cost_matrix[path[-1], path[0]])
            checked += 1
            if total_cost < best_cost:
                best_cost = float(total_cost)
                best_route = path[:]
            return None

        last = path[-1]
        candidates = sorted(unvisited, key=lambda j: cost_matrix[last, j])

        for nxt in candidates:
            edge_cost = float(cost_matrix[last, nxt])
            result = dfs(path + [nxt], unvisited - {nxt}, current_cost + edge_cost)
            if result == "TIME_LIMIT":
                return "TIME_LIMIT"
        return None

    if closed_loop:
        start = 0
        dfs([start], set(range(1, n)), 0.0)
    else:
        start_order = list(range(n))
        start_order.sort(key=lambda i: min(cost_matrix[i, j] for j in range(n) if j != i))
        for start in start_order:
            if t0 is not None and not within_limit(t0):
                break
            result = dfs([start], set(range(n)) - {start}, 0.0)
            if result == "TIME_LIMIT":
                break

    if t0 is not None and not within_limit(t0):
        if best_route is None:
            return None, None, f"TIME_LIMIT ({checked} complete routes checked)"
        return best_route, best_cost, f"TIME_LIMIT ({checked} complete routes checked)"

    return best_route, best_cost, "OPTIMAL"


def solve_gurobi_exact(
    cost_matrix: np.ndarray,
    closed_loop: bool,
    t0: Optional[float] = None,
    gurobi_license_config: Optional[Dict[str, Any]] = None,
):
    n = len(cost_matrix)

    if n == 0:
        return [], 0.0, "OPTIMAL"
    if n == 1:
        return [0], 0.0, "OPTIMAL"

    if not GUROBI_AVAILABLE:
        return None, None, "GUROBI_NOT_AVAILABLE"

    env = None
    model = None
    try:
        env = create_gurobi_env(gurobi_license_config)
        if env is None:
            model = gp.Model("changeover_tsp")
        else:
            model = gp.Model("changeover_tsp", env=env)
        model.Params.OutputFlag = 0

        rem = remaining_time_seconds(t0)
        if rem is not None:
            model.Params.TimeLimit = max(1.0, rem)

        if closed_loop:
            C = np.array(cost_matrix, dtype=float)
            N = n
            dummy = None
        else:
            N = n + 1
            dummy = n
            C = np.zeros((N, N), dtype=float)
            C[:n, :n] = cost_matrix
            C[dummy, :n] = 0.0
            C[:n, dummy] = 0.0
            C[dummy, dummy] = 0.0

        nodes = list(range(N))
        real_nodes = list(range(n))

        x = {}
        for i in nodes:
            for j in nodes:
                if i != j:
                    x[i, j] = model.addVar(vtype=GRB.BINARY, obj=float(C[i, j]), name=f"x_{i}_{j}")

        u = {i: model.addVar(lb=0.0, ub=N - 1, vtype=GRB.CONTINUOUS, name=f"u_{i}") for i in nodes}

        model.ModelSense = GRB.MINIMIZE
        model.update()

        for i in nodes:
            model.addConstr(gp.quicksum(x[i, j] for j in nodes if j != i) == 1)
        for j in nodes:
            model.addConstr(gp.quicksum(x[i, j] for i in nodes if i != j) == 1)

        model.addConstr(u[0] == 0)
        for i in nodes:
            if i != 0:
                model.addConstr(u[i] >= 1)
                model.addConstr(u[i] <= N - 1)

        for i in nodes:
            for j in nodes:
                if i != j and i != 0 and j != 0:
                    model.addConstr(u[i] - u[j] + N * x[i, j] <= N - 1)

        model.optimize()

        if model.SolCount == 0:
            if model.Status == GRB.TIME_LIMIT:
                return None, None, "TIME_LIMIT_NO_SOLUTION"
            return None, None, f"NO_SOLUTION_STATUS_{model.Status}"

        succ = {}
        for i in nodes:
            for j in nodes:
                if i != j and x[i, j].X > 0.5:
                    succ[i] = j
                    break

        if closed_loop:
            full_cycle = []
            cur = 0
            seen = set()
            while cur not in seen:
                seen.add(cur)
                full_cycle.append(cur)
                cur = succ[cur]

            route = [v for v in full_cycle if v in real_nodes]
            if len(route) != n:
                return None, None, "FAILED_BAD_ROUTE"
            total_cost = route_cost(route, cost_matrix, closed_loop=True)
        else:
            full_cycle = []
            cur = 0
            seen = set()
            while cur not in seen:
                seen.add(cur)
                full_cycle.append(cur)
                cur = succ[cur]

            if dummy not in full_cycle:
                return None, None, "FAILED_DUMMY_NOT_FOUND"

            k = full_cycle.index(dummy)
            rotated = full_cycle[k + 1:] + full_cycle[:k]
            route = [v for v in rotated if v != dummy]
            if len(route) != n:
                return None, None, "FAILED_BAD_ROUTE"
            total_cost = route_cost(route, cost_matrix, closed_loop=False)

        status = "OPTIMAL" if model.Status == GRB.OPTIMAL else "FEASIBLE"
        return route, total_cost, status

    except Exception as exc:
        return None, None, f"GUROBI_ERROR: {exc}"
    finally:
        try:
            if model is not None:
                model.dispose()
        except Exception:
            pass
        try:
            if env is not None:
                env.dispose()
        except Exception:
            pass


def solve_ortools_tsp(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    n = len(cost_matrix)

    if n == 0:
        return [], 0.0, "OPTIMAL"
    if n == 1:
        return [0], 0.0, "OPTIMAL"

    if not ORTOOLS_AVAILABLE:
        return None, None, "ORTOOLS_NOT_AVAILABLE"

    try:
        scale = 100000
        scaled = np.rint(np.array(cost_matrix, dtype=float) * scale).astype(np.int64)

        if closed_loop:
            N = n
            C = scaled
            manager = pywrapcp.RoutingIndexManager(N, 1, 0)
        else:
            N = n + 1
            dummy = n
            C = np.zeros((N, N), dtype=np.int64)
            C[:n, :n] = scaled
            C[dummy, :n] = 0
            C[:n, dummy] = 0
            C[dummy, dummy] = 0
            manager = pywrapcp.RoutingIndexManager(N, 1, [dummy], [dummy])

        routing = pywrapcp.RoutingModel(manager)

        def distance_callback(from_index, to_index):
            from_node = manager.IndexToNode(from_index)
            to_node = manager.IndexToNode(to_index)
            return int(C[from_node, to_node])

        transit_callback_index = routing.RegisterTransitCallback(distance_callback)
        routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

        search_parameters = pywrapcp.DefaultRoutingSearchParameters()
        search_parameters.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
        search_parameters.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH

        rem = remaining_time_seconds(t0)
        search_parameters.time_limit.seconds = 30 if rem is None else max(1, int(rem))

        solution = routing.SolveWithParameters(search_parameters)
        if solution is None:
            return None, None, "NO_SOLUTION"

        index = routing.Start(0)
        full_route = []
        while not routing.IsEnd(index):
            node = manager.IndexToNode(index)
            full_route.append(node)
            index = solution.Value(routing.NextVar(index))

        if closed_loop:
            route = full_route
            if len(route) != n:
                return None, None, "FAILED_BAD_ROUTE"
            total_cost = route_cost(route, cost_matrix, closed_loop=True)
        else:
            dummy = n
            route = [node for node in full_route if node != dummy]
            if len(route) != n:
                return None, None, "FAILED_BAD_ROUTE"
            total_cost = route_cost(route, cost_matrix, closed_loop=False)

        return route, total_cost, "FEASIBLE"

    except Exception as exc:
        return None, None, f"ORTOOLS_ERROR: {exc}"


# ------------------------------
# Heuristics
# ------------------------------
def solve_greedy(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    n = len(cost_matrix)
    best_route = None
    best_cost = math.inf

    for s in range(n):
        if t0 is not None and not within_limit(t0):
            break
        route = nearest_neighbor_route(cost_matrix, start=s)
        c = route_cost(route, cost_matrix, closed_loop)
        if c < best_cost:
            best_cost = c
            best_route = route

    if best_route is None:
        return None, None, "TIME_LIMIT_NO_SOLUTION"
    return best_route, best_cost, "FEASIBLE"


def minimum_spanning_tree_prim(sym: np.ndarray, t0: Optional[float] = None):
    import heapq

    n = len(sym)
    if n == 0:
        return []

    seen = [False] * n
    seen[0] = True
    heap = []
    tree = []

    for v in range(1, n):
        heapq.heappush(heap, (sym[0, v], 0, v))

    while heap and len(tree) < n - 1:
        if t0 is not None and not within_limit(t0):
            break
        w, u, v = heapq.heappop(heap)
        if seen[v]:
            continue
        seen[v] = True
        tree.append((u, v, w))
        for x in range(n):
            if not seen[x] and x != v:
                heapq.heappush(heap, (sym[v, x], v, x))

    return tree


def euler_tour_from_multigraph(adj_multi: Dict[int, List[int]], start: int = 0) -> List[int]:
    stack = [start]
    circuit = []
    local_adj = {k: v[:] for k, v in adj_multi.items()}

    while stack:
        v = stack[-1]
        if local_adj[v]:
            u = local_adj[v].pop()
            local_adj[u].remove(v)
            stack.append(u)
        else:
            circuit.append(stack.pop())

    circuit.reverse()
    return circuit


def minimum_matching_greedy(odd_vertices: List[int], sym: np.ndarray):
    unmatched = set(odd_vertices)
    pairs = []
    while unmatched:
        u = unmatched.pop()
        v = min(unmatched, key=lambda x: sym[u, x])
        unmatched.remove(v)
        pairs.append((u, v, sym[u, v]))
    return pairs


def solve_ant_colony(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    n = len(cost_matrix)
    if n <= 1:
        return list(range(n)), 0.0, "OPTIMAL"

    eta = 1.0 / np.maximum(cost_matrix, 1e-9)
    np.fill_diagonal(eta, 0.0)
    tau = np.ones((n, n), dtype=float)

    ants = min(30, max(8, n))
    alpha = 1.0
    beta = 3.0
    rho = 0.2
    q = 100.0

    best_route = None
    best_cost = math.inf

    iterations = 200 if TIME_LIMIT_SECONDS is None else max(20, min(500, int(TIME_LIMIT_SECONDS * 5)))

    for _ in range(iterations):
        if t0 is not None and not within_limit(t0):
            break

        ant_routes = []
        ant_costs = []

        for _a in range(ants):
            start = np.random.randint(0, n)
            unvisited = set(range(n))
            unvisited.remove(start)
            route = [start]
            cur = start

            while unvisited:
                candidates = list(unvisited)
                desirability = np.array(
                    [(tau[cur, j] ** alpha) * (eta[cur, j] ** beta) for j in candidates],
                    dtype=float,
                )
                s = desirability.sum()
                if s <= 0:
                    nxt = np.random.choice(candidates)
                else:
                    probs = desirability / s
                    nxt = np.random.choice(candidates, p=probs)

                route.append(nxt)
                unvisited.remove(nxt)
                cur = nxt

            route = two_opt(route, cost_matrix, closed_loop, t0=t0)
            c = route_cost(route, cost_matrix, closed_loop)

            ant_routes.append(route)
            ant_costs.append(c)

            if c < best_cost:
                best_cost = c
                best_route = route[:]

        tau *= (1 - rho)
        for route, c in zip(ant_routes, ant_costs):
            dep = q / max(c, 1e-9)
            for i in range(len(route) - 1):
                tau[route[i], route[i + 1]] += dep
            if closed_loop and len(route) > 1:
                tau[route[-1], route[0]] += dep

    if best_route is None:
        return None, None, "FAILED"

    return best_route, best_cost, "FEASIBLE"


def solve_christofides(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    n = len(cost_matrix)
    if n <= 2:
        route = list(range(n))
        return route, route_cost(route, cost_matrix, closed_loop), "FEASIBLE"

    sym = symmetrize_matrix(cost_matrix, mode="avg")
    mst = minimum_spanning_tree_prim(sym, t0=t0)
    if len(mst) != n - 1:
        return None, None, "FAILED"

    degree = [0] * n
    for u, v, _ in mst:
        degree[u] += 1
        degree[v] += 1
    odd = [i for i in range(n) if degree[i] % 2 == 1]

    matching = minimum_matching_greedy(odd, sym)

    adj_multi = {i: [] for i in range(n)}
    for u, v, _ in mst + matching:
        adj_multi[u].append(v)
        adj_multi[v].append(u)

    tour = euler_tour_from_multigraph(adj_multi, start=0)

    seen = set()
    route = []
    for x in tour:
        if x not in seen:
            seen.add(x)
            route.append(x)

    if len(route) != n:
        return None, None, "FAILED"

    route = two_opt(route, cost_matrix, closed_loop, t0=t0)
    return route, route_cost(route, cost_matrix, closed_loop), "FEASIBLE"


def solve_two_opt_three_opt_combined(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    route, _, status = solve_greedy(cost_matrix, closed_loop, t0=t0)
    if route is None:
        return None, None, status
    route = two_opt(route, cost_matrix, closed_loop, t0=t0)
    route = three_opt_simple(route, cost_matrix, closed_loop, t0=t0)
    return route, route_cost(route, cost_matrix, closed_loop), "FEASIBLE"


def solve_lin_kernighan_lite(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    route, _, status = solve_greedy(cost_matrix, closed_loop, t0=t0)
    if route is None:
        return None, None, status

    best = route[:]
    best_cost = route_cost(best, cost_matrix, closed_loop)

    improved = True
    while improved:
        improved = False

        cand = two_opt(best, cost_matrix, closed_loop, t0=t0)
        cand_cost = route_cost(cand, cost_matrix, closed_loop)
        if cand_cost + 1e-12 < best_cost:
            best, best_cost = cand, cand_cost
            improved = True

        cand = three_opt_simple(best, cost_matrix, closed_loop, t0=t0)
        cand_cost = route_cost(cand, cost_matrix, closed_loop)
        if cand_cost + 1e-12 < best_cost:
            best, best_cost = cand, cand_cost
            improved = True

        if t0 is not None and not within_limit(t0):
            break

    return best, best_cost, "FEASIBLE"


def solve_simulated_annealing(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    n = len(cost_matrix)
    if n == 0:
        return [], 0.0, "OPTIMAL"
    if n == 1:
        return [0], 0.0, "OPTIMAL"

    route = list(np.random.permutation(n))
    route = two_opt(route, cost_matrix, closed_loop, t0=t0)
    current_cost = route_cost(route, cost_matrix, closed_loop)

    best = route[:]
    best_cost = current_cost

    T = 1000.0
    cooling = 0.995

    while T > 1:
        if t0 is not None and not within_limit(t0):
            break

        i, j = sorted(np.random.choice(range(n), 2, replace=False))
        new_route = route[:]
        new_route[i:j] = reversed(new_route[i:j])
        new_cost = route_cost(new_route, cost_matrix, closed_loop)

        if new_cost < current_cost or np.random.rand() < np.exp((current_cost - new_cost) / max(T, 1e-9)):
            route = new_route
            current_cost = new_cost
            if new_cost < best_cost:
                best = new_route[:]
                best_cost = new_cost

        T *= cooling

    return best, best_cost, "FEASIBLE"


def solve_tabu_search(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    n = len(cost_matrix)
    if n == 0:
        return [], 0.0, "OPTIMAL"
    if n == 1:
        return [0], 0.0, "OPTIMAL"

    route = nearest_neighbor_route(cost_matrix)
    route = two_opt(route, cost_matrix, closed_loop, t0=t0)
    best = route[:]
    best_cost = route_cost(best, cost_matrix, closed_loop)

    tabu_list = []
    tenure = 10

    for _ in range(200):
        if t0 is not None and not within_limit(t0):
            break

        best_candidate_route = None
        best_candidate_cost = math.inf
        best_move = None

        for i in range(1, n - 1):
            for j in range(i + 1, n):
                move = (i, j)
                if move in tabu_list:
                    continue

                new_route = route[:]
                new_route[i:j] = reversed(new_route[i:j])
                c = route_cost(new_route, cost_matrix, closed_loop)

                if c < best_candidate_cost:
                    best_candidate_route = new_route
                    best_candidate_cost = c
                    best_move = move

        if best_candidate_route is None:
            break

        route = best_candidate_route
        tabu_list.append(best_move)
        if len(tabu_list) > tenure:
            tabu_list.pop(0)

        if best_candidate_cost < best_cost:
            best = route[:]
            best_cost = best_candidate_cost

    return best, best_cost, "FEASIBLE"


def solve_cheapest_insertion(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    n = len(cost_matrix)
    if n == 0:
        return [], 0.0, "OPTIMAL"
    if n == 1:
        return [0], 0.0, "OPTIMAL"
    if n == 2:
        route = [0, 1]
        return route, route_cost(route, cost_matrix, closed_loop), "FEASIBLE"

    route = [0, 1]

    for k in range(2, n):
        if t0 is not None and not within_limit(t0):
            break

        best_pos = None
        best_increase = float("inf")

        if closed_loop:
            m = len(route)
            for i in range(m):
                j = (i + 1) % m
                increase = (
                    float(cost_matrix[route[i]][k])
                    + float(cost_matrix[k][route[j]])
                    - float(cost_matrix[route[i]][route[j]])
                )
                if increase < best_increase:
                    best_increase = increase
                    best_pos = j
        else:
            candidates = []

            increase_start = float(cost_matrix[k][route[0]])
            candidates.append((increase_start, 0))

            increase_end = float(cost_matrix[route[-1]][k])
            candidates.append((increase_end, len(route)))

            for j in range(1, len(route)):
                i = j - 1
                increase = (
                    float(cost_matrix[route[i]][k])
                    + float(cost_matrix[k][route[j]])
                    - float(cost_matrix[route[i]][route[j]])
                )
                candidates.append((increase, j))

            best_increase, best_pos = min(candidates, key=lambda x: x[0])

        route.insert(best_pos, k)

    return route, route_cost(route, cost_matrix, closed_loop), "FEASIBLE"


def solve_tsp_metaheuristic(cost_matrix: np.ndarray, closed_loop: bool, t0: Optional[float] = None):
    n = len(cost_matrix)
    if n == 0:
        return [], 0.0, "OPTIMAL"
    if n == 1:
        return [0], 0.0, "OPTIMAL"

    best_route = None
    best_cost = math.inf

    for s in range(n):
        if t0 is not None and not within_limit(t0):
            break
        route = nearest_neighbor_route(cost_matrix, start=s)
        route = two_opt(route, cost_matrix, closed_loop, t0=t0)
        c = route_cost(route, cost_matrix, closed_loop)
        if c < best_cost:
            best_route, best_cost = route[:], c

    restarts = max(30, min(200, 10 * n))
    for _ in range(restarts):
        if t0 is not None and not within_limit(t0):
            break
        route = list(np.random.permutation(n))
        route = two_opt(route, cost_matrix, closed_loop, t0=t0)
        if n <= 60:
            route = three_opt_simple(route, cost_matrix, closed_loop, t0=t0)
        c = route_cost(route, cost_matrix, closed_loop)
        if c < best_cost:
            best_route, best_cost = route[:], c

    if best_route is None:
        return None, None, "FAILED"

    return best_route, best_cost, "FEASIBLE"


SOLVER_MAP = {
    FULL_ENUMERATION_LABEL: solve_full_enumeration,
    "Gurobi Exact": solve_gurobi_exact,
    "OR-Tools TSP": solve_ortools_tsp,
    "Ant Colony Optimization": solve_ant_colony,
    "Christofides Algorithm": solve_christofides,
    "2-opt / 3-opt": solve_two_opt_three_opt_combined,
    "Lin-Kernighan": solve_lin_kernighan_lite,
    "Simulated Annealing": solve_simulated_annealing,
    "Tabu Search": solve_tabu_search,
    "Cheapest Insertion": solve_cheapest_insertion,
    "TSP Heuristic": solve_tsp_metaheuristic,
}


def choose_best_result(results: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    valid = [
        r for r in results
        if r["total_changeover"] is not None
        and isinstance(r["total_changeover"], (int, float, np.floating))
        and np.isfinite(r["total_changeover"])
        and r["route"] is not None
    ]
    if not valid:
        return None
    return min(valid, key=lambda r: (r["total_changeover"], r["runtime_seconds"]))


def run_selected_methods(
    display_labels: List[str],
    cost_matrix: np.ndarray,
    closed_loop: bool,
    selected_algorithms: List[str],
    progress_callback=None,
    gurobi_license_config: Optional[Dict[str, Any]] = None,
) -> Tuple[List[Dict[str, Any]], Optional[Dict[str, Any]]]:
    n = len(display_labels)
    original_time_limit_seconds = TIME_LIMIT_SECONDS

    if n <= 15:
        algorithms_to_run = [FULL_ENUMERATION_LABEL]
    else:
        algorithms_to_run = [algo for algo in selected_algorithms if algo in SOLVER_MAP]

    results = []
    total_algorithms = len(algorithms_to_run)

    for idx, algo_name in enumerate(algorithms_to_run, start=1):
        if progress_callback is not None:
            try:
                progress_callback(idx, total_algorithms, algo_name)
            except Exception:
                pass

        if original_time_limit_seconds is None:
            per_algorithm_limit_seconds = None
        else:
            remaining_algorithms = max(1, total_algorithms - idx + 1)
            used_seconds = sum(item["runtime_seconds"] for item in results)
            remaining_total_seconds = max(1.0, float(original_time_limit_seconds) - float(used_seconds))
            per_algorithm_limit_seconds = max(1, int(remaining_total_seconds / remaining_algorithms))

        refresh_time_limit_seconds(
            None if per_algorithm_limit_seconds is None else per_algorithm_limit_seconds / 3600.0
        )

        solver = SOLVER_MAP[algo_name]
        t0 = now()
        try:
            if algo_name == "Gurobi Exact":
                route, total_cost, status = solver(
                    cost_matrix,
                    closed_loop,
                    t0,
                    gurobi_license_config=gurobi_license_config,
                )
            else:
                route, total_cost, status = solver(cost_matrix, closed_loop, t0)
        except Exception as exc:
            route, total_cost, status = None, None, f"ERROR: {exc}"
        runtime = round(elapsed(t0), 4)
        display_sequence = route_to_display_sequence(route, display_labels, closed_loop=closed_loop)
        results.append(
            {
                "algorithm": algo_name,
                "status": status,
                "route": route,
                "sequence": display_sequence,
                "sequence_text": " → ".join(display_sequence) if display_sequence else "-",
                "total_changeover": None if total_cost is None else float(total_cost),
                "runtime_seconds": runtime,
            }
        )

    refresh_time_limit_seconds(None if original_time_limit_seconds is None else original_time_limit_seconds / 3600.0)
    best_result = choose_best_result(results)
    return results, best_result


def build_output_dataframe(results: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for idx, item in enumerate(results, start=1):
        total_changeover = item["total_changeover"]
        rows.append(
            {
                "No.": idx,
                "Algorithm": item["algorithm"],
                "Total Changeover Time": "-" if total_changeover is None else round(float(total_changeover), 4),
                "Runtime": f'{round(float(item["runtime_seconds"]), 4)} s',
                "Optimal Sequence": item["sequence_text"] if item["sequence_text"] else "-",
            }
        )
    return pd.DataFrame(rows)


def _update_input_sheet(workbook, selected_display_products: List[str], problem_type: str, time_limit_hours: Any) -> None:
    if "Input Data" not in workbook.sheetnames:
        return

    ws = workbook["Input Data"]

    for row in range(1, 501):
        ws.cell(row=row, column=1, value=None)

    ws["A1"] = "Product"
    for idx, product in enumerate(selected_display_products, start=2):
        ws.cell(row=idx, column=1, value=product)

    ws["C1"] = "Problem"
    ws["D1"] = problem_type
    ws["C2"] = "Time Limit (Hours)"
    ws["D2"] = time_limit_hours


def write_results_to_excel(
    matrix_path: Path | str,
    output_excel_path: Path | str,
    selected_display_products: List[str],
    problem_type: str,
    time_limit_hours: Any,
    results: List[Dict[str, Any]],
    best_result: Optional[Dict[str, Any]],
    closed_loop: bool,
) -> Path:
    matrix_path = Path(matrix_path)
    output_excel_path = Path(output_excel_path)

    shutil.copyfile(matrix_path, output_excel_path)
    wb = load_workbook(output_excel_path)

    _update_input_sheet(wb, selected_display_products, problem_type, time_limit_hours)

    if "Output" in wb.sheetnames:
        ws = wb["Output"]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet("Output")

    ws["A1"] = "Optimal Sequence No."
    ws["B1"] = "Product"
    ws["C1"] = "Transition"

    if best_result is not None and best_result["sequence"]:
        best_sequence = best_result["sequence"]
        transitions = []
        for i in range(len(best_sequence) - 1):
            transitions.append(f"{best_sequence[i]} -> {best_sequence[i + 1]}")
        for idx, product in enumerate(best_sequence, start=1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=product)
        for idx, transition in enumerate(transitions, start=2):
            ws.cell(row=idx, column=3, value=transition)

    table_headers = ["No.", "Algorithm", "Total Changeover Time", "Runtime", "Optimal Sequence"]
    start_col = 5
    ws.cell(row=1, column=start_col, value="Comparison Results")
    for offset, header in enumerate(table_headers):
        ws.cell(row=2, column=start_col + offset, value=header)

    output_df = build_output_dataframe(results)
    for row_idx, row in enumerate(output_df.itertuples(index=False), start=3):
        ws.cell(row=row_idx, column=start_col, value=row[0])
        ws.cell(row=row_idx, column=start_col + 1, value=row[1])
        ws.cell(row=row_idx, column=start_col + 2, value=row[2] if row[2] != "-" else None)
        ws.cell(row=row_idx, column=start_col + 3, value=row[3])
        ws.cell(row=row_idx, column=start_col + 4, value=row[4])

    summary_row = max(15, len(results) + 6)
    ws.cell(row=summary_row, column=5, value="Best Method")
    ws.cell(row=summary_row, column=6, value=(best_result["algorithm"] if best_result else "None"))
    ws.cell(row=summary_row + 1, column=5, value="Minimum Changeover Time")
    ws.cell(
        row=summary_row + 1,
        column=6,
        value=(round(float(best_result["total_changeover"]), 4) if best_result and best_result["total_changeover"] is not None else None),
    )
    ws.cell(row=summary_row + 2, column=5, value="Loop Type")
    ws.cell(row=summary_row + 2, column=6, value=problem_type)
    ws.cell(row=summary_row + 3, column=5, value="Time Limit (Hours)")
    ws.cell(row=summary_row + 3, column=6, value=str(time_limit_hours))

    wb.save(output_excel_path)
    return output_excel_path


def create_output_workbook_path(matrix_path: Path | str) -> Path:
    matrix_path = Path(matrix_path)
    return matrix_path.with_name(f"{matrix_path.stem}_with_results.xlsx")


def summarize_statuses(results: List[Dict[str, Any]]) -> str:
    failed = [r for r in results if r["route"] is None or r["total_changeover"] is None]
    if not failed:
        return ""
    parts = [f'{item["algorithm"]}: {item["status"]}' for item in failed]
    return " Some selected algorithms did not return a usable route: " + "; ".join(parts) + "."


def run_optimizer(
    selected_display_products: List[str],
    problem_type: str,
    time_limit_hours: Any,
    selected_algorithms: List[str],
    matrix_excel_path: Path | str,
    progress_callback=None,
    gurobi_license_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    if not selected_display_products:
        raise ValueError("Please select at least one product.")

    display_labels, cost_matrix, _actual_labels = get_cost_submatrix(selected_display_products, matrix_excel_path)
    closed_loop = normalize_loop_type(problem_type) == "closed"
    refresh_time_limit_seconds(time_limit_hours)
    normalized_gurobi_license_config = normalize_gurobi_license_config(gurobi_license_config)

    results, best_result = run_selected_methods(
        display_labels=display_labels,
        cost_matrix=cost_matrix,
        closed_loop=closed_loop,
        selected_algorithms=selected_algorithms,
        progress_callback=progress_callback,
        gurobi_license_config=normalized_gurobi_license_config,
    )

    if not results:
        raise ValueError("No algorithm was available to run.")

    output_excel_path = create_output_workbook_path(matrix_excel_path)
    write_results_to_excel(
        matrix_path=matrix_excel_path,
        output_excel_path=output_excel_path,
        selected_display_products=selected_display_products,
        problem_type=problem_type,
        time_limit_hours=time_limit_hours,
        results=results,
        best_result=best_result,
        closed_loop=closed_loop,
    )

    output_df = build_output_dataframe(results)
    csv_data = output_df.to_csv(index=False).encode("utf-8")
    excel_bytes = output_excel_path.read_bytes()

    if best_result is None:
        sequence = []
        sequence_text = "-"
        total_changeover = None
        runtime_seconds = round(max(item["runtime_seconds"] for item in results), 4)
        best_algorithm = None
        solver_note = "No selected algorithm produced a usable route." + summarize_statuses(results)
    else:
        sequence = best_result["sequence"]
        sequence_text = best_result["sequence_text"]
        total_changeover = round(float(best_result["total_changeover"]), 4)
        runtime_seconds = round(float(best_result["runtime_seconds"]), 4)
        best_algorithm = best_result["algorithm"]
        if len(selected_display_products) <= 15:
            solver_note = (
                "15 or fewer products were selected, so the run was fixed to Enumeration."
                + summarize_statuses(results)
            )
        else:
            solver_note = (
                "The best feasible result among the selected algorithms is shown below."
                + summarize_statuses(results)
            )

    return {
        "status": "Completed",
        "problem_type": "Close" if closed_loop else "Open",
        "time_limit_hours": time_limit_hours,
        "selected_count": len(selected_display_products),
        "selected_algorithms": [item["algorithm"] for item in results],
        "sequence": sequence,
        "sequence_text": sequence_text,
        "algorithm_results": results,
        "best_algorithm": best_algorithm,
        "total_changeover": total_changeover,
        "runtime_seconds": runtime_seconds,
        "solver_note": solver_note,
        "output_df": output_df,
        "csv_data": csv_data,
        "output_excel_path": str(output_excel_path),
        "output_excel_name": output_excel_path.name,
        "output_excel_bytes": excel_bytes,
    }
